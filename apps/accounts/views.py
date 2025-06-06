""" Vistas de la aplicación de cuentas de usuario. """
import openpyxl
import mimetypes
from functools import reduce
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.urls import reverse
from django.views.decorators.http import require_POST
from formtools.wizard.views import SessionWizardView
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Value, Case, When
from django.middleware.csrf import get_token
from django.utils import timezone
from django.core.mail import send_mail
from apps.notifications.models import Notification
from apps.supplier.forms import CatSupplierForm, SupplierDocumentForm
from apps.supplier.models import CatSupplier, Supplier, SupplierDocument, DocumentType
from apps.accounts.forms import StaffUserCreationForm, StaffUserEditForm, SupplierUserCreationForm, SupplierUserEditForm, ProfileUpdateForm, CustomSupplierCreationForm, UserForm, SupplierFiscalForm, SupplierBankForm, UserProfileForm
from apps.accounts.models import Profile


DOCUMENT_FIELDS = [
    'fiscal_situation', 'address_proof', 'bank_statement', 'sat_positive_opinion', 'incorporation_act', 'legal_power', 'identification', 'imss_opinion', 'infonavit_opinion', 'supplier_registration_request', 'repse', 'confidentiality_agreement', 'induction_course'
]

# Create your views here.


def register(request):
    """Vista para el registro de usuarios."""
    if request.method == 'POST':
        form = CustomSupplierCreationForm(request.POST)
        if form.is_valid():
            # Guardamos el usuario en la tabla de Users
            user = form.save()
            # user.id
            # Vamos a verificar si el RFC ya se encuentra en el modelo CatSupplier
            catSupplier, created = CatSupplier.objects.get_or_create(
                rfc=form.cleaned_data['rfc'],
                defaults={
                    'software_origin': 'Aplicativo',
                    'company': 'Aplicativo (Todas las empresas)',
                    'external_id': 0,
                    'name': form.cleaned_data['company'],
                    'email': user.email,
                    'active': False,
                    'created_at': timezone.now(),
                }
            )

            # Crear instancia de Supplier asociada al usuario
            supplier = Supplier(
                user=user,
                company_name=form.cleaned_data['company'],
                rfc=form.cleaned_data['rfc'].upper(),
                address=form.cleaned_data['address'],
                contact_person=form.cleaned_data['contact_person'],
                phone_number=form.cleaned_data['phone_number'],
                supplier_type=form.cleaned_data['supplier_type'],
                tax_regime=form.cleaned_data['tax_regime'],
                email=user.email,
            )
            supplier.save()

            # Crear notificación de bienvenida
            Notification.objects.create(
                created_by=user,
                recipient=user,
                title='Bienvenido al Portal de Proveedores de TotalGas',
                description='Por favor, llena tu información para poder ser considerado como proveedor de TotalGas.',
                link='/supplier/profile',
                notification_type='info',
                priority=10,
                expires_at=timezone.now()+timezone.timedelta(days=7),
                action_text='Llenar Información'
            )

            messages.success(
                request, f'¡El usuario {user.username} fue creado correctamente!')

            email = form.cleaned_data.get('email')
            if email:
                # send_custom_email(request, 'Confirmación de registro en el Portal de Proveedores de TotalGas', 'registro_proveedor.html', {
                #     'proveedor_nombre': form.cleaned_data['contact_person']
                # }, [email])
                send_verification_email(supplier)

            accept_supplier_url = request.build_absolute_uri(
                reverse('accept_supplier', args=[user.id])
            )
            docs_validation_url = request.build_absolute_uri(
                reverse('docs_validation', args=[])
            )
            # Enviar correo de confirmación
            context = {
                'username': user.username,
                'company': form.cleaned_data['company'],
                'email': user.email,
                'phone_number': form.cleaned_data['phone_number'],
                'rfc': form.cleaned_data['rfc'],
                'contact_person': form.cleaned_data['contact_person'],
                'address': form.cleaned_data['address'],
                'supplier_type': form.cleaned_data['supplier_type'],
                'tax_regime': form.cleaned_data['tax_regime'],
                'accept_supplier_url': accept_supplier_url,
                'docs_validation_url': docs_validation_url,
            }

            # Vamos a obtener los correos electrónicos de todos los usuarios que son administradores
            recipients = User.objects.filter(
                is_staff=True).values_list('email', flat=True)
            # Vasmoa a poner los correos en una lista separada por comas
            recipients = list(recipients)
            # Vamos a guardar la ruta accept_supplier en una variable
            send_custom_email(request, 'Solicitud de Alta en Portal de Proveedores', 'alta_proveedor.html', context, recipients)

            # Redirigimos al usuario a la página de inicio de sesión
            return redirect('login')
        else:
            # Si el formulario tiene errores, se recarga con los datos ingresados
            return render(request, 'accounts/register.html', {'form': form})
    else:
        form = CustomSupplierCreationForm()

    return render(request, 'accounts/register.html', {'form': form})

def send_verification_email(supplier):
    subject = "Bienvenido al Portal de Proveedores"
    context = {
        'company_name': supplier.company_name,
        'verification_token': supplier.verification_token,
        'verification_url': f"{settings.SITE_URL}{reverse('verify_supplier', args=[supplier.verification_token])}"
    }

    send_mail(
        subject,
        (
            f"Antes de activar tu cuenta, es necesario que cargues los documentos requeridos en el siguiente enlace: {context['verification_url']}\n\n"
            "Una vez que hayas completado este paso, un administrador revisará tu información y activará tu cuenta.\n\n"
            "Si tienes alguna duda, no dudes en contactarnos."
        ),
        settings.DEFAULT_FROM_EMAIL,
        [supplier.email],
        fail_silently=False,
    )



def login_view(request):
    """Vista para el login de usuarios."""
    # Capturar el next de GET o POST
    next_url = request.POST.get('next') or request.GET.get('next', '')

    if request.method == 'POST':
        username_or_rfc = request.POST.get('username') # Capturamos el campo "Usuario o RFC"
        password = request.POST.get('password')

        # Buscar el usuario usando el RFC o nombre de usuario
        try:
            user = User.objects.get(username=username_or_rfc)  # Intentamos buscar por nombre de usuario
        except User.DoesNotExist:
            user = None

        if user is not None and user.check_password(password):  # Verificamos que las contraseñas coincidan
            # Verificar si el usuario está activo
            if not user.is_active:
                messages.error(
                    request,
                    '<h3 style="text-align: center; color: #FF3333; padding-top: 25px">Tu cuenta no está activada. Contacta con el administrador.</h3>'
                )
                return render(request, 'accounts/login.html', {'next': next_url})

            login(request, user)

            # Redirigir a next_url si está presente
            if next_url:
                return redirect(next_url)

            # Redirigir según el tipo de usuario
            if user.is_staff:
                messages.success(
                    request, f"¡Bienvenido de nuevo, {user.username} (Admin)!")
                return redirect('administrator_index')
            else:
                messages.success(
                    request, f"¡Bienvenido de nuevo, {user.username} (Proveedor)!")
                return redirect('supplier_index')
        else:
            # Manejo de error si las credenciales son incorrectas
            messages.error(
                request,
                '<h3 style="text-align: center; color: #FF3333; padding-top: 25px">Usuario o contraseña incorrectos. Intente de nuevo.</h3>'
            )
            return render(request, 'accounts/login.html', {
                'username': username_or_rfc,
                'next': next_url  # Mantener el next en caso de error
            })

    # Renderizar el formulario de login con next_url
    return render(request, 'accounts/login.html', {'next': next_url})


@login_required
def show(request):
    """Vista para mostrar los usuarios del sistema."""
    # Aqui vamos a recuperar los datos de la base de datos de usuarios
    staff_users = User.objects.filter(is_staff=True)

    # Vamos a pasar el formulario al template
    staff_form = StaffUserCreationForm()

    return render(request, 'accounts/show.html', {'staff_users': staff_users, 'staff_form': staff_form})


@login_required
def create_staff_user(request):
    """Vista para crear un usuario del sistema."""
    if request.method == 'POST':
        # Vamos a recuperar el username
        staff_form = StaffUserCreationForm(request.POST)
        if staff_form.is_valid():
            staff_form.save()  # Guardar el usuario
            messages.success(
                request, f"¡El usuario {staff_form.cleaned_data.get('username')} fue creado correctamente!")
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            # Aqui vamos a recuperar los datos de la base de datos de usuarios
            messages.error(
                request, f'Por favor, corrija los siguientes errores: {staff_form.errors}')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        return HttpResponse('GET request')
        form = StaffUserCreationForm()

    return redirect('show')


@login_required
def create_supplier_user(request):
    """Vista para crear un usuario del sistema."""
    if request.method == 'POST':
        # Vamos a recuperar el username
        supplier_form = SupplierUserCreationForm(request.POST)
        if supplier_form.is_valid():
            supplier_form.save()  # Guardar el usuario
            messages.success(
                request, f"¡El usuario {supplier_form.cleaned_data.get('username')} fue creado correctamente!")
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            # Aqui vamos a recuperar los datos de la base de datos de usuarios
            messages.error(
                request, f'Por favor, corrija los siguientes errores: {supplier_form.errors}')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))

    return redirect('show')


@login_required
def delete_account(request, user_id):
    """Vista para eliminar un usuario del sistema."""
    formtype = request.POST.get('formtype')

    # Aqui vamos a recuperar el usuario
    user = User.objects.get(id=user_id)
    user.delete()

    messages.success(
        request, f'¡El usuario {user.username} fue eliminado correctamente!')
    if formtype == 'staff':
        return redirect('/accounts/show')
    else:
        return redirect('/accounts/show_suppliers')


@login_required
def password_change_modal(request, user_id):
    """Vista para cambiar la contraseña de un usuario."""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = PasswordChangeForm(user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'¡La contraseña del usuario {user.username} fue cambiada correctamente!')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.error(
                request, f'Se encontrarón los siguientes errores: {form.errors}')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        form = PasswordChangeForm(user)
        return JsonResponse({'success': True, 'html': render(request, 'accounts/modals/password_change_modal.html', {'form': form}).content.decode('utf-8')})


@login_required
def profile_modal(request, user_id):
    """Vista para mostrar el modal de perfil de usuario."""
    user = get_object_or_404(User, id=user_id)
    profile, _ = Profile.objects.get_or_create(user=user)

    if request.method == 'POST':
        form = UserProfileForm(
            request.POST, request.FILES, instance=user, profile_instance=profile
        )
        if form.is_valid():
            form.save(profile_instance=profile)
            messages.success(
                request,
                f'¡Los datos del usuario {user.username} han sido actualizados correctamente!'
            )
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            return JsonResponse({
                'success': False,
                'html': render(request, 'supplier/modals/profile_modal.html', {'form': form}).content.decode('utf-8')
            })

    form = UserProfileForm(instance=user, profile_instance=profile)
    return JsonResponse({
        'success': True,
        'html': render(request, 'supplier/modals/profile_modal.html', {'form': form}).content.decode('utf-8')
    })



@login_required
def update_staff_account(request, user_id):
    """Vista para actualizar los datos de un usuario."""
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = StaffUserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'¡Los datos del usuario {user.username} han sido actualizados correctamente!')
            return redirect('/accounts/show')
        else:
            # Si el formulario no es válido, puede ser útil mostrar los errores
            messages.error(
                request, f'Se encontrarón los siguientes errores: {form.errors}')
            return redirect('/accounts/show')
    else:
        form = StaffUserEditForm(instance=user)
        return JsonResponse({'success': True, 'html': render(request, 'accounts/modals/update_staff_account.html', {'form': form, 'user_id': user_id}).content.decode('utf-8')})


@login_required
def update_supplier_account(request, user_id):
    """Vista para actualizar los datos de un usuario."""
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = SupplierUserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'¡Los datos del usuario {user.username} han sido actualizados correctamente!')
            return redirect('/accounts/show_suppliers')
        else:
            # Si el formulario no es válido, puede ser útil mostrar los errores
            messages.error(
                request, f'Se encontrarón los siguientes errores: {form.errors}')
            return redirect('/accounts/show_suppliers')
    else:
        form = SupplierUserEditForm(instance=user)
        return JsonResponse({'success': True, 'html': render(request, 'accounts/modals/update_supplier_account.html', {'form': form, 'user_id': user_id}).content.decode('utf-8')})


@login_required
def show_suppliers(request):
    """Vista para mostrar los proveedores del sistema."""
    # Vamos a pasar el formulario al template
    supplier_form = SupplierUserCreationForm()

    return render(request, 'accounts/show_suppliers.html', {'supplier_form': supplier_form})


def show_suppliers_table(request):
    """Vista para mostrar los proveedores del sistema."""
    # Aqui vamos a recuperar los datos de la base de datos de usuarios
    suppliers_users = User.objects.filter(is_staff=False)
    csrf_token = get_token(request)

    # 3️⃣ Crear la lista de datos para la respuesta JSON
    data = [
        {
            'id': supplier.id,
            'username': supplier.username,
            'name': supplier.first_name + ' ' + supplier.last_name,
            'superuser': ('<span class="badge bg-success">Activo</span>' if supplier.is_superuser else '<span class="badge bg-danger">Inactivo</span>'),
            'email': supplier.email,
            'is_active': ('<span class="badge bg-success">Activo</span>' if supplier.is_active else '<span class="badge bg-danger">Inactivo</span>'),
            'date_joined': supplier.date_joined,
            'last_login': supplier.last_login,
            'actions': f'''
                <div class="btn-group btn-group-sm">
                    <button type="button" class="btn btn-info dropdown-toggle" data-bs-toggle="dropdown" aria-haspopup="true" aria-expanded="false">Acciones</button>
                    <div class="dropdown-menu">
                    <a class="dropdown-item" href="javascript:void(0);" data-bs-toggle="modal" data-bs-target="#updateSupplierModal" data-id="{supplier.id}"><i class="mdi mdi-lead-pencil"></i> Editar</a>
                    <form method="POST" action="/accounts/delete_account/{supplier.id}/" class="dropdown-item">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <button type="submit" class="btn btn-link text-danger p-0 m-0" style="border: none; background: none;">Eliminar</button>
                    </form>
                    </div>
                </div>
            ''',
        }
        for supplier in suppliers_users
    ]
    # 4️⃣ Devolver los datos como JSON
    return JsonResponse({'data': data})


@login_required
def show_catsuppliers(request):
    """Vista para mostrar los proveedores del sistema."""
    print(request.method)
    if request.method == 'GET':
        # Lógica para solicitud GET
        return render(request, 'accounts/show_catsuppliers.html')


from django.db.models import Count, Q

@login_required
def show_table_catsuppliers(request):
    """Vista para mostrar los proveedores del sistema."""

    # 1️⃣ Obtener códigos de documentos requeridos
    REQUIRED_DOCUMENT_CODES = list(
        DocumentType.objects.filter(is_required=True).values_list('code', flat=True)
    )

    # 2️⃣ Anotar los documentos subidos requeridos por proveedor
    suppliers = Supplier.objects.annotate(
        total_documents=Count(
            'documents',
            filter=Q(documents__document_type__code__in=REQUIRED_DOCUMENT_CODES)
                   & ~Q(documents__file=None)
                   & ~Q(documents__file=''),
            distinct=True
        )
    ).values('rfc', 'total_documents')

    # Convertir a diccionario para acceso rápido por RFC
    supplier_docs = {supplier['rfc']: supplier['total_documents']
                     for supplier in suppliers}

    # 3️⃣ Consultar todos los CatSupplier
    cat_suppliers = CatSupplier.objects.all()
    csrf_token = get_token(request)

    # 4️⃣ Generar los datos del response
    data = [
        {
            "id": cat_supplier.id,
            "name": cat_supplier.name,
            "rfc": cat_supplier.rfc,
            "postal_code": cat_supplier.postal_code,
            "city": cat_supplier.city,
            "state": cat_supplier.state,
            "email": cat_supplier.email,
            "bank": cat_supplier.bank,
            "account_number": cat_supplier.account_number,
            "clabe": cat_supplier.clabe,
            "payment_method": cat_supplier.payment_method,
            "currency": cat_supplier.currency,
            "website": cat_supplier.website,
            "category": cat_supplier.category,
            "active": ('<span class="badge bg-success">Activo</span>' if cat_supplier.active else '<span class="badge bg-danger">Inactivo</span>'),
            "notes": cat_supplier.notes,
            # 👈 Número de documentos requeridos subidos por proveedor
            "digital_file": supplier_docs.get(cat_supplier.rfc, 0),
            'actions': f'''
                <div class="btn-group btn-group-sm">
                    <button type="button" class="btn btn-info dropdown-toggle" data-bs-toggle="dropdown" aria-haspopup="true" aria-expanded="false">Acciones</button>
                    <div class="dropdown-menu">
                    <a class="dropdown-item" href="javascript:void(0);" data-bs-toggle="modal" data-bs-target="#update_cat_supplier_modal" data-id="{cat_supplier.id}">Editar</a>
                    <a class="dropdown-item" href="/accounts/update_digital_files/{cat_supplier.id}/" target="_blank">Actualizar archivo</a>
                    <form method="POST" action="/accounts/delete_catsupplier/{cat_supplier.id}/" class="dropdown-item">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <button type="submit" class="btn btn-link text-danger p-0 m-0" style="border: none; background: none;">Eliminar</button>
                    </form>
                    </div>
                </div>
            ''',
        }
        for cat_supplier in cat_suppliers
    ]

    return JsonResponse({'data': data})



def update_cat_supplier_modal(request, supplier_id):
    """Vista para mostrar el modal de actualización de proveedor."""
    supplier = get_object_or_404(CatSupplier, id=supplier_id)

    if request.method == 'POST':
        form = CatSupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'¡Los datos del proveedor {supplier.name} han sido actualizados correctamente!')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
        else:
            messages.error(
                request, f'Se encontrarón los siguientes errores: {form.errors}')
            # '/' es la página de fallback
            return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        form = CatSupplierForm(instance=supplier)
        return JsonResponse({
            'success': True,
            'html': render(request, 'supplier/modals/update_catsupplier.html', {'form': form, 'supplier': supplier}).content.decode('utf-8')
        })


@login_required
def delete_catsupplier(request, supplier_id):
    """Vista para eliminar un proveedor del sistema."""
    supplier = CatSupplier.objects.get(id=supplier_id)
    supplier.delete()

    messages.success(
        request, f'¡El proveedor {supplier.name} fue eliminado correctamente!')
    return redirect('/accounts/show_catsuppliers')


@login_required
def update_digital_files(request, supplier_id):
    """Vista para actualizar los archivos digitales de un proveedor."""
    # si la peticion es GET
    if request.method == 'GET':
        # Vamos a buscar al proveedor en la tabla CatSupplier
        catSupplier = get_object_or_404(CatSupplier, id=supplier_id)

        # Verificamos si existe un registro en el modelo Supplier
        if not Supplier.objects.filter(rfc=catSupplier.rfc).exists():
            # Si no existe, redirigimos a la vista de creación
            return HttpResponse(f'El proveedor {catSupplier.name} no tiene un registro de usuario asociado.')

        # Si existe, redirigimos a la vista de actualización
        supplier = Supplier.objects.get(rfc=catSupplier.rfc)
        return render(request, 'supplier/update_digital_files.html', {'supplier': supplier})
    else:
        user = get_object_or_404(User, id=request.POST.get('user_id'))

        try:
            # Intenta obtener el proveedor asociado al usuario
            supplier = Supplier.objects.get(user=user)
        except Supplier.DoesNotExist:
            # Redirige al formulario de creación si no existe el proveedor
            return redirect('supplier_create', user_id=user.id)

        # Si el proveedor existe, permite editar
        # form = SupplierForm(request.POST, request.FILES, instance=supplier)
        # if form.is_valid():
        #     form.save()
        #     messages.success(
        #         request, f'¡El proveedor {form.cleaned_data.get('company_name')} fue actualizado correctamente!')
        #     return redirect(request.META.get('HTTP_REFERER', '/'))
        # else:
        #     messages.error(
        #         request, f'Se encontrarón los siguientes errores: {form.errors}')
        #     return redirect(request.META.get('HTTP_REFERER', '/'))


def send_custom_email(request, subject, template, context, recipients):
    """Función para enviar correos electrónicos personalizados."""
    # Renderizar el template HTML que hemos creado
    message = render_to_string(f'emails/{template}', context)

    # Enviar el correo
    send_mail(
        subject,
        '',  # El cuerpo en texto plano si lo quieres
        'TotalGas - Portal de Proveedores <totalgasdesarrollo@gmail.com>',  # El remitente
        recipients,  # El destinatario
        html_message=message,  # El mensaje en HTML
    )


def accept_supplier(request, user_id):
    """Vista para aceptar un proveedor en el sistema."""
    user = get_object_or_404(User, id=user_id)
    user.is_active = True
    user.save()
    supplier = Supplier.objects.get(user=user)

    messages.success(
        request, f'¡El proveedor {supplier.company_name} ha sido activado correctamente!')
    return render(request, 'supplier/accept_supplier.html', {'supplier': supplier})


def staffusers_table(request):
    """Vista para mostrar los usuarios del sistema."""
    # Aqui vamos a recuperar los datos de la base de datos de usuarios
    staff_users = User.objects.filter(is_staff=True)
    csrf_token = get_token(request)
    # 3️⃣ Crear la lista de datos para la respuesta JSON
    data = [
        {
            'id': user.id,
            'username': user.username.lower(),
            'name': user.first_name + ' ' + user.last_name,
            'superuser': ('<span class="badge bg-success">Activo</span>' if user.is_superuser else '<span class="badge bg-danger">Inactivo</span>'),
            'email': user.email.lower(),
            'status': ('<span class="badge bg-success">Activo</span>' if user.is_active else '<span class="badge bg-danger">Inactivo</span>'),
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Nunca ha ingresado',
            'actions': f'''
                <div class="btn-group btn-group-sm">
                    <button type="button" class="btn btn-info dropdown-toggle" data-bs-toggle="dropdown" aria-haspopup="true" aria-expanded="false">Acciones</button>
                    <div class="dropdown-menu">
                    <a class="dropdown-item" href="javascript:void(0);" data-bs-toggle="modal" data-bs-target="#update_cat_supplier_modal" data-id="{user.id}">Editar</a>
                    <a class="dropdown-item" href="/accounts/update_digital_files/{user.id}/" target="_blank">Actualizar archivo</a>
                    <form method="POST" action="/accounts/delete_catsupplier/{user.id}/" class="dropdown-item">
                        <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                        <button type="submit" class="btn btn-link text-danger p-0 m-0" style="border: none; background: none;">Eliminar</button>
                    </form>
                    </div>
                </div>
            ''',
        }
        for user in staff_users
    ]

    # 4️⃣ Devolver los datos como JSON
    return JsonResponse({'data': data})


def export_suppliers_to_excel(request):
    """Vista para exportar los proveedores a un archivo de Excel."""
    # Crear un libro de Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Definir los encabezados de la tabla
    headers = [
        "ID", "Nombre", "RFC", "Código Postal", "Ciudad", "Estado", "Correo", "Banco", "No de Cuenta", "CLABE", "Método de Pago", "Moneda", "Sitio Web", "Categoría", "Notas"
    ]
    ws.append(headers)

    # Consultar los datos desde la base de datos (ajusta según tu modelo)
    suppliers = CatSupplier.objects.all().values_list(
        "id", "name", "rfc", "postal_code", "city", "state", "email", "bank", "account_number", "clabe", "payment_method", "currency", "website", "category", "notes"
    )

    # Agregar los datos al Excel
    for supplier in suppliers:
        ws.append(supplier)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="ProveedoresLayout.xlsx"'

    # Guardar el archivo en la respuesta
    wb.save(response)
    return response


def import_suppliers_from_excel(request):
    """Vista para importar proveedores desde un archivo de Excel."""
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]

        # Cargar el archivo Excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Iterar sobre las filas del Excel (asumiendo que los encabezados están en la primera fila)
            # Empieza en la segunda fila
            for row in ws.iter_rows(min_row=2, values_only=True):
                supplier_data = {
                    "name": row[1],
                    "rfc": row[2],
                    "postal_code": row[3],
                    "city": row[4],
                    "state": row[5],
                    "email": row[6],
                    "bank": row[7],
                    "account_number": row[8],
                    "clabe": row[9],
                    "payment_method": row[10],
                    "currency": row[11],
                    "website": row[12],
                    "category": row[13],
                    "notes": row[14],
                }

                print(supplier_data)

                # Actualizar el proveedor correspondiente al id de la columna A (row[0])
                supplier, created = CatSupplier.objects.update_or_create(
                    id=row[0],  # Usamos el id de la columna A
                    defaults=supplier_data  # Los datos a actualizar
                )

            messages.success(
                request, '¡Los proveedores han sido importados correctamente!')
            return redirect('show_catsuppliers')

        except Exception as e:
            messages.error(request, f'Error al importar los proveedores: {e}')
            return redirect('show_catsuppliers')

    return HttpResponse("Método no permitido", status=405)


class SupplierWizard(SessionWizardView):
    """Vista para el registro de proveedores en un wizard."""
    form_list = [UserForm, SupplierFiscalForm, SupplierBankForm]
    template_name = "accounts/register2.html"

    def done(self, form_list, **kwargs):
        user_form, fiscal_form, bank_form = form_list

        # Guardar usuario
        user = user_form.save(commit=False)
        user.set_password(user_form.cleaned_data['password'])
        user.save()

        # Crear perfil de proveedor
        supplier = fiscal_form.save(commit=False)
        supplier.user = user  # Relacionar el proveedor con el usuario creado
        supplier.save()

        # Guardar información bancaria
        bank_data = bank_form.cleaned_data
        Supplier.objects.filter(id=supplier.id).update(**bank_data)

        # Vamos a crear una notificacion de bienvenida para el proveedor
        Notification.objects.create(
            recipient=user,
            title='Bienvenido al Portal de Proveedores de TotalGas',
            description='Por favor, llena tu información para poder ser considerado como proveedor de TotalGas.',
            link='/supplier/profile',
            notification_type='info',
            priority=10,
            expires_at=timezone.now() + timezone.timedelta(days=7),
            action_text='Llenar Información'
        )

        messages.success(
            self.request, f'¡El usuario {user.username} fue creado correctamente!')

        # Enviar correo de confirmación al usuario
        send_custom_email(self.request, 'Confirmación de registro en el Portal de Proveedores de TotalGas',
                          'registro_proveedor.html', {'proveedor_nombre': user.first_name}, [user.email])

        # Enviar correo de confirmación al administrador
        accept_supplier_url = self.request.build_absolute_uri(
            f'/accounts/accept_supplier/{user.id}/')
        context = {
            'username': user.username,
            'company': supplier.company_name,
            'email': user.email,
            'phone_number': supplier.phone_number,
            'rfc': supplier.rfc,
            'contact_person': supplier.contact_person,
            'address': supplier.address,
            'supplier_type': supplier.supplier_type,
            'tax_regime': supplier.tax_regime,
        }

        # Vamos a obtener los correos electrónicos de todos los usuarios que son administradores
        recipients = User.objects.filter(
            is_staff=True).values_list('email', flat=True)
        # Vasmos a poner los correos en una lista separada por comas
        recipients = list(recipients)

        # Vamos a guardar la ruta accept_supplier en una variable
        send_custom_email(self.request, 'Solicitud de Alta en Portal de Proveedores',
                          'alta_proveedor.html', context, recipients)

        return redirect('login')  # Redirigir a una página de éxito

def verify_supplier(request, verification_token):
    supplier = get_object_or_404(Supplier, verification_token=verification_token)
    supplier.verified = True
    supplier.save()
    messages.success(request, '¡Tu cuenta ha sido verificada correctamente!')
    # Vamos a llevar al usuario a un portal de registro de documentos llamado upload_docs, pasando el id del proveedor
    return redirect(reverse('upload_docs', kwargs={'verification_token': supplier.verification_token}))

@require_POST
def upload_document(request, supplier_id, document_type_id):
    """Vista para cargar o actualizar un documento individual"""
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    document_type = get_object_or_404(DocumentType, pk=document_type_id)
    
    # Buscar si ya existe un documento de este tipo
    document = SupplierDocument.objects.filter(
        supplier=supplier,
        document_type=document_type
    ).first()
    
    # Creamos uno nuevo
    form = SupplierDocumentForm(
        request.POST,
        request.FILES,
        document_type=document_type,
        supplier=supplier,
        instance=document
    )
    
    # Validar el formulario
    if form.is_valid():
        document = form.save()
        
        # Actualizar el estado del documento
        document.status = 'pending'
        document.save()

        # Enviar correo de confirmación al administrador
        user = supplier.user
        docs_validation_url = request.build_absolute_uri(reverse('docs_validation', args=[]))

        context = {
            'username': user.username,
            'company': supplier.company_name,
            'email': user.email,
            'phone_number': supplier.phone_number,
            'rfc': supplier.rfc,
            'contact_person': supplier.contact_person,
            'address': supplier.address,
            'supplier_type': supplier.supplier_type,
            'tax_regime': supplier.tax_regime,
            'document_type': document_type.name,
            'docs_validation_url': docs_validation_url,
        }

        # Vamos a obtener los correos electrónicos de todos los usuarios que son administradores
        recipients = User.objects.filter(
            is_staff=True).values_list('email', flat=True)
        # Vasmos a poner los correos en una lista separada por comas
        recipients = list(recipients)

        # Vamos a guardar la ruta accept_supplier en una variable
        send_custom_email(request, f"El documento '{document_type.name}' ha sido actualizado", 'documento_actualizado.html', context, recipients)

        messages.success(request, f"Documento '{document_type.name}' cargado correctamente.")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'document_status': document.status,
                'uploaded_at': document.uploaded_at.strftime('%d/%m/%Y %H:%M'),
                'message': f"Documento '{document_type.name}' cargado correctamente."
            })
    else:
        error_message = "Error al cargar el documento. Por favor, verifique el archivo."
        messages.error(request, error_message)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'errors': form.errors,
                'message': error_message
            }, status=400)
    
    return redirect(reverse('upload_docs', kwargs={'verification_token': supplier.verification_token}))

@require_POST
def update_document_status(request, document_id):
    """Vista para actualizar el estado de un documento (aprobar/rechazar)"""
    document = get_object_or_404(SupplierDocument, pk=document_id)
    supplier = get_object_or_404(Supplier, pk=document.supplier.id)
    
    # Verificar permisos (aquí podrías agregar lógica para verificar si el usuario tiene permisos)
    
    status = request.POST.get('status')
    rejection_reason = request.POST.get('rejection_reason', '')
    
    if status in ['approved', 'rejected']:
        document.status = status
        document.reviewed_at = timezone.now()
        
        if status == 'rejected':
            document.rejection_reason = rejection_reason
        
        document.save()
        
        message = "Documento aprobado correctamente." if status == 'approved' else "Documento rechazado correctamente."
        messages.success(request, message)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'document_status': document.status,
                'reviewed_at': document.reviewed_at.strftime('%d/%m/%Y %H:%M'),
                'message': message
            })
    else:
        messages.error(request, "Estado no válido.")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': "Estado no válido."
            }, status=400)
    
    return redirect(reverse('upload_docs', kwargs={'verification_token': supplier.verification_token}))

def delete_document(request, document_id):
    """Vista para eliminar un documento"""
    document = get_object_or_404(SupplierDocument, pk=document_id)
    supplier = get_object_or_404(Supplier, pk=document.supplier.id)
    
    # Verificar permisos
    document_name = document.document_type.name
    document.delete()
    
    messages.success(request, f"Documento '{document_name}' eliminado correctamente.")
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'message': f"Documento '{document_name}' eliminado correctamente."
        })
    
    return redirect(reverse('upload_docs', kwargs={'verification_token': supplier.verification_token}))

def upload_docs(request, verification_token):
    # Obtener el proveedor usando el supplier_id de la URL
    supplier = get_object_or_404(Supplier, verification_token=verification_token)

    document_types = DocumentType.objects.all().order_by('name')
    
    # Preparar los datos para cada tipo de documento
    documents_data = []
    
    for doc_type in document_types:
        # Buscar si existe un documento para este tipo
        document = SupplierDocument.objects.filter(
            supplier=supplier,
            document_type=doc_type
        ).first()
        
        # Crear un formulario para este tipo de documento
        form = SupplierDocumentForm(
            document_type=doc_type,
            supplier=supplier,
            instance=document
        )
        
        documents_data.append({
            'type': doc_type,
            'document': document,
            'form': form,
        })
    
    return render(request, 'accounts/upload_docs.html', {
        'supplier': supplier,
        'documents_data': documents_data,
    })

def view_document(request, document_id):
    document = get_object_or_404(SupplierDocument, id=document_id)
    
    # Determinar el tipo de contenido basado en la extensión del archivo
    content_type, encoding = mimetypes.guess_type(document.file.name)
    if not content_type:
        content_type = 'application/octet-stream'
    
    # Archivos que se pueden visualizar en navegador
    viewable_types = ['application/pdf', 'image/jpeg', 'image/png', 'image/gif', 'text/plain']
    
    if content_type in viewable_types:
        # Servir para visualización
        response = FileResponse(document.file.open('rb'), content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{document.file.name.split("/")[-1]}"'
    else:
        # Servir para descarga
        response = FileResponse(document.file.open('rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{document.file.name.split("/")[-1]}"'
    
    return response
